import requests
import openpyxl
import time
from datetime import time as dt_time
from pathlib import Path
from openpyxl.styles import PatternFill

API_KEY = 'SEM_VLOZ_GOOGLE_API_KEY'

INPUT_FILE = 'Data collection test zákazníci.xlsx'
OUTPUT_FILE = 'Data collection test zákazníci_doplneno.xlsx'

COL_NAME = 2
COL_ADDR1 = 3
COL_PSC = 4
COL_CITY = 5

COL_OPEN_FROM_1 = 7
COL_OPEN_TO_1 = 8
COL_OPEN_FROM_2 = 9
COL_OPEN_TO_2 = 10

COL_PLACE_ID = 32
COL_GOOGLE_NAME = 33
COL_GOOGLE_ADDRESS = 34
COL_LATITUDE = 35
COL_LONGITUDE = 36

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

GOOGLE_INDEX_TO_OFFSET = {
    1: 0,
    2: 1,
    3: 2,
    4: 3,
    5: 4,
    6: 5,
    0: 6,
}

print('FULL all-in-one Opening Hours version stored in repository')
print('Google Places API -> Place ID -> Opening Hours -> Latitude -> Longitude -> Excel export')
